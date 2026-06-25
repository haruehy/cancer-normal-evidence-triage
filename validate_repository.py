#!/usr/bin/env python3
"""Lightweight repository validation for manuscript reproducibility files.

This check verifies the bundled metadata, supplementary tables, curated
annotations, and processed outputs without requiring large raw public datasets.
Full regeneration of all analyses is handled by run_all.py after raw PRISM,
DepMap, ChEMBL, and NCI-ALMANAC files have been downloaded into data/.
"""

from __future__ import annotations

import csv
import json
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent

REQUIRED_PATHS = [
    "README.md",
    "requirements.txt",
    "run_all.py",
    "analysis/config.py",
    "analysis/01_build_compound_annotations.py",
    "data/README.md",
    "data/curated_compound_annotations.csv",
    "outputs/tables/compound_annotation_counts.csv",
    "outputs/tables/prism_secondary_repair_vs_removal_group_summary.csv",
    "outputs/tables/cell_line_RTR_scores_with_quadrants_and_variants.csv",
    "outputs/tables/cross_dataset_drug_summary.csv",
    "outputs/tables/cross_dataset_axis_summary.csv",
    "outputs/tables/primary_logfc_candidate_axis_results.csv",
    "outputs/tables/secondary_auc_candidate_axis_results.csv",
    "outputs/tables/candidate_axis_dependency_screen.csv",
    "outputs/tables/candidate_gene_dependency_screen.csv",
    "outputs/tables/RTR_weight_sensitivity_top100_overlap.csv",
    "tables/S6_sensitivity_analyses/RTR_weight_sensitivity_top100_overlap.csv",
    "tables/S4_normal_protection_evidence_literature_csv/S4_PN_half_point_sensitivity_anchor_candidates.csv",
    "tables/S3_Table_ChEMBL_normal_like_assay_audit.xlsx",
    "tables/S4_Table_normal_protection_evidence_literature.xlsx",
    "tables/S14_input_file_provenance_manifest.csv",
]

EXPECTED_COUNTS = {
    "strict repair/protective": 6,
    "broad repair/protective": 75,
    "cytotoxic/removal-like": 157,
    "other": 1261,
}


def fail(message: str) -> None:
    raise SystemExit(f"VALIDATION FAILED: {message}")


def check_required_paths() -> None:
    missing = [p for p in REQUIRED_PATHS if not (ROOT / p).exists()]
    if missing:
        fail("missing required paths: " + ", ".join(missing))
    print("[ok] required files and folders are present")


def check_curated_annotation_counts() -> None:
    ann = pd.read_csv(ROOT / "data/curated_compound_annotations.csv")
    if "analysis_group" not in ann.columns:
        fail("data/curated_compound_annotations.csv must contain analysis_group")
    counts = ann.groupby("analysis_group").size().to_dict()
    for group, expected in EXPECTED_COUNTS.items():
        observed = int(counts.get(group, 0))
        if observed != expected:
            fail(f"analysis_group count mismatch for {group}: observed {observed}, expected {expected}")
    print("[ok] curated compound annotation counts match manuscript values")


def check_supplementary_tables() -> None:
    for rel in [
        "tables/S3_Table_ChEMBL_normal_like_assay_audit.xlsx",
        "tables/S4_Table_normal_protection_evidence_literature.xlsx",
    ]:
        wb = load_workbook(ROOT / rel, read_only=True, data_only=True)
        if not wb.sheetnames:
            fail(f"{rel} has no sheets")
    s14 = pd.read_csv(ROOT / "tables/S14_input_file_provenance_manifest.csv")
    required_s14_cols = {"source_resource", "filename", "release_or_access_label"}
    if not required_s14_cols.issubset(set(s14.columns)):
        fail("S14 manifest is missing expected provenance columns")
    print("[ok] S3/S4 workbooks and S14 manifest are readable")


def check_all_supplementary_present() -> None:
    """Every manuscript supplement S1..S14 must have an explicit S-numbered deliverable."""
    tables = ROOT / "tables"
    missing = []
    for n in range(1, 15):
        prefix = f"S{n}_"
        if not any(p.name.startswith(prefix) for p in tables.iterdir()):
            missing.append(f"S{n}")
    if missing:
        fail("missing explicit supplementary deliverables for: " + ", ".join(missing))
    if not (tables / "REPOSITORY_TABLES_INDEX.csv").exists():
        fail("tables/REPOSITORY_TABLES_INDEX.csv is missing")
    print("[ok] all repository tables S1-S14 have explicit deliverables")


def check_processed_outputs() -> None:
    counts = pd.read_csv(ROOT / "outputs/tables/compound_annotation_counts.csv")
    group_col = "analysis_group"
    if group_col not in counts.columns:
        fail("outputs/tables/compound_annotation_counts.csv must contain analysis_group")
    if not counts.get("matches_manuscript", pd.Series(dtype=bool)).astype(bool).all():
        fail("compound_annotation_counts.csv reports at least one mismatch")
    key_outputs = list((ROOT / "outputs/tables").glob("*.csv"))
    if len(key_outputs) < 10:
        fail("too few processed output tables were found")
    print("[ok] processed output tables are present and internally consistent")


def check_lineage_adjusted_dependency_regression_gene_set() -> None:
    required_genes = {"ATM", "XIAP"}
    csv_paths = [
        ROOT / "outputs/tables/lineage_adjusted_dependency_regression.csv",
        ROOT / "tables/S11_depmap_crispr_rescreen/lineage_adjusted_dependency_regression.csv",
    ]
    frames = []
    for path in csv_paths:
        if not path.exists():
            fail(f"missing lineage-adjusted dependency regression file: {path.relative_to(ROOT)}")
        df = pd.read_csv(path)
        if "comparison" not in df.columns or "gene" not in df.columns:
            fail(f"{path.relative_to(ROOT)} must contain comparison and gene columns")
        for comparison, sub in df.groupby("comparison"):
            genes = set(sub["gene"].astype(str))
            missing = sorted(required_genes - genes)
            if missing:
                fail(
                    f"{path.relative_to(ROOT)} is missing required lineage-adjusted genes "
                    f"for {comparison}: " + ", ".join(missing)
                )
        frames.append((path, df))

    left_path, left = frames[0]
    right_path, right = frames[1]
    sort_cols = [c for c in ["comparison", "gene", "axis"] if c in left.columns and c in right.columns]
    left_cmp = left.sort_values(sort_cols).reset_index(drop=True) if sort_cols else left.reset_index(drop=True)
    right_cmp = right.sort_values(sort_cols).reset_index(drop=True) if sort_cols else right.reset_index(drop=True)
    if list(left_cmp.columns) != list(right_cmp.columns) or not left_cmp.equals(right_cmp):
        fail(
            "lineage_adjusted_dependency_regression.csv differs between "
            f"{left_path.relative_to(ROOT)} and {right_path.relative_to(ROOT)}"
        )

    xlsx_path = ROOT / "tables/S11_Table_depmap_crispr_rescreen.xlsx"
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.startswith("lineage_adjusted_dependency_reg")), None)
    if sheet_name is None:
        fail("S11 workbook is missing lineage_adjusted_dependency_reg sheet")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        fail("S11 workbook lineage-adjusted sheet is empty")
    header = [str(v) if v is not None else "" for v in rows[0]]
    if "comparison" not in header or "gene" not in header:
        fail("S11 workbook lineage-adjusted sheet must contain comparison and gene columns")
    cidx = header.index("comparison")
    gidx = header.index("gene")
    by_comparison = {}
    for row in rows[1:]:
        if row is None or len(row) <= max(cidx, gidx) or row[cidx] is None or row[gidx] is None:
            continue
        by_comparison.setdefault(str(row[cidx]), set()).add(str(row[gidx]))
    for comparison, genes in by_comparison.items():
        missing = sorted(required_genes - genes)
        if missing:
            fail(
                "S11 workbook lineage-adjusted sheet is missing required genes "
                f"for {comparison}: " + ", ".join(missing)
            )
    if len(rows) - 1 != len(left):
        fail(
            "S11 workbook lineage-adjusted sheet row count does not match "
            "outputs/tables/lineage_adjusted_dependency_regression.csv"
        )
    print("[ok] lineage-adjusted dependency regression includes ATM/XIAP and synchronized S11 copies")


def check_script_smoke_test() -> None:
    # 01_build_compound_annotations uses the bundled curated table only and should run without raw public data.
    result = subprocess.run(
        [sys.executable, str(ROOT / "analysis/01_build_compound_annotations.py")],
        cwd=ROOT,
        text=True,
        capture_output=True,
    )
    if result.returncode != 0:
        print(result.stdout)
        print(result.stderr, file=sys.stderr)
        fail("01_build_compound_annotations.py smoke test failed")
    expected = ROOT / "outputs/tables/secondary_drug_compound_annotations.csv"
    if not expected.exists():
        fail("smoke test did not generate secondary_drug_compound_annotations.csv")
    print("[ok] smoke test for curated compound annotation script passed")


    result9 = subprocess.run(
        [sys.executable, str(ROOT / "analysis/09_generate_tables_and_figures.py")],
        cwd=ROOT,
        text=True,
        capture_output=True,
    )
    if result9.returncode != 0:
        print(result9.stdout)
        print(result9.stderr, file=sys.stderr)
        fail("09_generate_tables_and_figures.py smoke test failed")
    print("[ok] smoke test for table/figure generation script passed")



def check_table3_table4_values() -> None:
    def close(observed: float, expected: float, tol: float = 5e-4) -> bool:
        return abs(float(observed) - float(expected)) <= tol

    gene = pd.read_csv(ROOT / "outputs/tables/candidate_gene_dependency_screen.csv")
    table3_expected = [
        ("RTR_high_removal_high_vs_other_no_Bone_SoftTissue", "KIF11", 0.032828, 0.524204, 0.498249, 0.900007),
        ("RTR_high_removal_high_vs_other_no_Bone_SoftTissue", "ATM", 0.092491, 0.634280, 0.000170, 0.016513),
        ("RTR_top20_vs_bottom20_no_Bone_SoftTissue", "XIAP", 0.062354, 0.659494, 0.000355, 0.017214),
        ("RTR_top20_vs_bottom20_no_Bone_SoftTissue", "MCL1", 0.219407, 0.607876, 0.015729, 0.177368),
        ("RTR_top20_vs_bottom20_no_Bone_SoftTissue", "TUBA1B", 0.176247, 0.607595, 0.016003, 0.177368),
    ]
    for comparison, gene_name, delta, mw_auc, p_value, fdr in table3_expected:
        row = gene[(gene["comparison"] == comparison) & (gene["gene"] == gene_name)]
        if row.empty:
            fail(f"missing Table 3 source row for {comparison} / {gene_name}")
        row = row.iloc[0]
        checks = [
            (row["delta"], delta, "delta"),
            (row["mw_auc"], mw_auc, "mw_auc"),
            (row["p"], p_value, "p"),
            (row["FDR_within_comparison"], fdr, "FDR_within_comparison"),
        ]
        for observed, expected, name in checks:
            if not close(observed, expected):
                fail(f"Table 3 source mismatch for {gene_name} {name}: observed {observed}, expected {expected}")

    sec = pd.read_csv(ROOT / "outputs/tables/secondary_auc_candidate_axis_results.csv")
    primary = pd.read_csv(ROOT / "outputs/tables/primary_logfc_candidate_axis_results.csv")
    table4_expected = {
        "apoptosis_IAP_BCL_MCL": (0.021073, 0.701130, 1.0453623859299445e-07, 0.131772, 0.657054, 4.954223960432397e-06),
        "KIF11_kinesin": (0.109273, 0.664614, 9.37192196831992e-06, 0.289628, 0.583751, 0.0155432817751789),
        "microtubule_tubulin": (0.034784, 0.664093, 9.37192196831992e-06, 0.528795, 0.624872, 0.0002308748631032),
        "DDR_ATR_CHK_WEE_PARP": (0.016364, 0.641878, 0.000121333602885, 0.221191, 0.671009, 1.031896952098224e-06),
        "PLK_Aurora_mitotic_kinase": (0.031218, 0.637936, 0.0001546847245818, 0.429432, 0.644446, 2.1851652534883808e-05),
        "HSP90_proteostasis": (0.021098, 0.628262, 0.0003848410251227, 0.233516, 0.578481, 0.0205199107146549),
    }
    comparison = "RTR_high_removal_high_vs_other_no_Bone_SoftTissue"
    for axis, expected in table4_expected.items():
        sec_row = sec[(sec["comparison"] == comparison) & (sec["axis"] == axis)]
        pri_row = primary[(primary["comparison"] == comparison) & (primary["axis"] == axis)]
        if sec_row.empty or pri_row.empty:
            fail(f"missing Table 4 source row for {axis}")
        sec_row = sec_row.iloc[0]
        pri_row = pri_row.iloc[0]
        observed_values = [
            sec_row["delta_sensitivity"], sec_row["mw_auc_sensitivity"], sec_row["FDR_within_comparison"],
            pri_row["delta_sensitivity"], pri_row["mw_auc_sensitivity"], pri_row["FDR_within_comparison"],
        ]
        for observed, expected_value in zip(observed_values, expected):
            if not close(observed, expected_value):
                fail(f"Table 4 source mismatch for {axis}: observed {observed}, expected {expected_value}")
    print("[ok] Table 3 and Table 4 source values match the reconciled manuscript values")


def check_sensitivity_summaries() -> None:
    rtr = pd.read_csv(ROOT / "outputs/tables/RTR_weight_sensitivity_top100_overlap.csv")
    expected = {
        ("default_combined_RTR_top100", "default", "default"): 100,
        ("repair_preservation_weighted_overlay", "0.7", "0.3"): 63,
        ("removal_vulnerability_weighted_overlay", "0.3", "0.7"): 78,
    }
    for (score_definition, rp, rv), expected_overlap in expected.items():
        row = rtr[
            (rtr["score_definition"].astype(str) == score_definition)
            & (rtr["repair_preservation_weight"].astype(str) == rp)
            & (rtr["removal_vulnerability_weight"].astype(str) == rv)
        ]
        if row.empty:
            fail(f"missing RTR weight sensitivity row: {score_definition} {rp}/{rv}")
        observed = int(row.iloc[0]["overlap_with_default_top100_n"])
        if observed != expected_overlap:
            fail(f"RTR weight sensitivity overlap mismatch for {score_definition}: observed {observed}, expected {expected_overlap}")

    pn = pd.read_csv(ROOT / "tables/S4_normal_protection_evidence_literature_csv/S4_PN_half_point_sensitivity_anchor_candidates.csv")
    if set(pn["compound"]) != {"Trolox", "L-ergothioneine", "Taurine", "Benfotiamine"}:
        fail("PN half-point sensitivity table does not contain the four anchor candidates")
    taurine = pn[pn["compound"] == "Taurine"].iloc[0]
    if taurine["plus_0p5_grade"] != "High" or taurine["minus_0p5_grade"] != "Moderate":
        fail("PN half-point sensitivity should flag Taurine as a Moderate/High boundary case")
    print("[ok] RTR weight and PN half-point sensitivity summaries match manuscript statements")

def check_rtr_decomposition_values() -> None:
    summ = pd.read_csv(ROOT / "outputs/tables/RTR_component_summary_and_correlations.csv")
    def comp(name):
        row = summ[summ["component"] == name]
        if row.empty:
            fail(f"RTR component summary missing {name}")
        return row.iloc[0]
    rp = comp("combined_repair_preservation_rank")
    rv = comp("combined_removal_vulnerability_rank")
    if int(rp["n"]) != 578 or int(rv["n"]) != 578:
        fail(f"RTR combined decomposition n mismatch: observed {int(rp['n'])}/{int(rv['n'])}, expected 578/578")
    if abs(float(rp["corr_with_RTR_spearman"]) - 0.361) > 1e-3:
        fail(f"RTR repair-preservation rho mismatch: observed {rp['corr_with_RTR_spearman']}, expected 0.361")
    if abs(float(rv["corr_with_RTR_spearman"]) - 0.679) > 1e-3:
        fail(f"RTR removal-vulnerability rho mismatch: observed {rv['corr_with_RTR_spearman']}, expected 0.679")
    print("[ok] RTR combined decomposition reproduces n=578 and rho 0.361/0.679")


def main() -> None:
    check_required_paths()
    check_curated_annotation_counts()
    check_supplementary_tables()
    check_all_supplementary_present()
    check_processed_outputs()
    check_lineage_adjusted_dependency_regression_gene_set()
    check_sensitivity_summaries()
    check_rtr_decomposition_values()
    check_script_smoke_test()
    check_table3_table4_values()
    print("\nRepository validation completed successfully.")
    print("Full regeneration with run_all.py requires raw public datasets listed in data/README.md and S14.")


if __name__ == "__main__":
    main()
